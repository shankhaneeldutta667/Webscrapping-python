from bs4 import BeautifulSoup
import requests
import openpyxl

excel=openpyxl.Workbook()
#print(excel.sheetnames)
sheet=excel.active
sheet.title='Imdb 250 shows'
#print(excel.sheetnames)
sheet.append(['Rank','Name','Year Released','Imdb Rating'])

try:
        source=requests.get('https://www.imdb.com/chart/top/')
        source.raise_for_status()
        soup=BeautifulSoup(source.text,'html.parser')
        movies=soup.find('tbody',class_='lister-list').find_all('tr')
        
        
        for movie in movies:
            exact_movie_name=movie.find('td',class_='titleColumn').a.text
            movie_details=movie.find('td',class_='titleColumn').get_text(strip=True)
            movie_rank=movie.find('td',class_='titleColumn').get_text(strip=True).split('.')[0]
            year=movie.find('td',class_='titleColumn').span.text.strip('()')
            rating=movie.find('td',class_='ratingColumn imdbRating').strong.text.strip(None)
            
            print(movie_rank,exact_movie_name,year,rating)
            sheet.append([movie_rank,exact_movie_name,year,rating])
            

except Exception as e:
    print(e)
excel.save('IMDB 250 Highest rated shows.xlsx')
